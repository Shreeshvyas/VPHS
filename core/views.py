import os
import shutil
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, Q
from django.http import HttpResponse, FileResponse, Http404
from django.utils.timezone import now
from django.conf import settings

# Excel export using openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from users.decorators import staff_only, admin_only, super_admin_only
from users.middleware import log_activity
from students.models import Student, StudentEnrollment
from staff.models import Staff, SalaryPayment
from fees.models import FeeCollection, StudentFeeStructure, ClassFeeStructure
from finance.models import Expense, Income, CashBookEntry

@login_required
@staff_only
def dashboard(request):
    today = now().date()
    current_month = today.month
    current_year = today.year
    
    # 1. Total counts
    total_students = Student.objects.filter(status='ACTIVE', is_deleted=False).count()
    total_staff = Staff.objects.filter(status='ACTIVE', is_deleted=False).count()
    
    # 2. Collections
    today_collections = FeeCollection.objects.filter(payment_date=today, is_deleted=False).aggregate(Sum('amount_paid'))['amount_paid__sum'] or 0.00
    
    start_of_month = today.replace(day=1)
    monthly_collections = FeeCollection.objects.filter(payment_date__gte=start_of_month, is_deleted=False).aggregate(Sum('amount_paid'))['amount_paid__sum'] or 0.00
    
    # 3. Pending fees
    pending_dues = StudentFeeStructure.objects.filter(is_deleted=False).exclude(status='PAID')
    total_pending = sum(d.balance_amount for d in pending_dues)
    
    # 4. Expenses
    monthly_expenses = Expense.objects.filter(
        expense_date__gte=start_of_month,
        status='APPROVED',
        is_deleted=False
    ).aggregate(Sum('amount'))['amount__sum'] or 0.00
    
    # 5. Ledger Balances
    latest_ledger = CashBookEntry.objects.filter(is_deleted=False).order_by('-entry_date', '-id').first()
    cash_in_hand = latest_ledger.running_cash_balance if latest_ledger else 0.00
    bank_balance = latest_ledger.running_bank_balance if latest_ledger else 0.00
    
    # 6. Recent transactions
    recent_transactions = CashBookEntry.objects.filter(is_deleted=False).order_by('-entry_date', '-id')[:6]
    
    # 7. Upcoming payroll
    upcoming_salaries = SalaryPayment.objects.filter(month=current_month, year=current_year, payment_status='PENDING', is_deleted=False).count()
    
    # 8. Charts data: Monthly cashflow (last 6 months)
    cashflow_chart = []
    months_labels = []
    for i in range(5, -1, -1):
        # Calculate month offsets
        m = (current_month - i - 1) % 12 + 1
        y = current_year - 1 if (current_month - i - 1) < 0 else current_year
        months_labels.append(f"{m}/{y}")
        
        # Debits (collections + income) in that month
        m_start = datetime.date(y, m, 1)
        if m == 12:
            m_end = datetime.date(y + 1, 1, 1) - datetime.timedelta(days=1)
        else:
            m_end = datetime.date(y, m + 1, 1) - datetime.timedelta(days=1)
            
        m_debits = CashBookEntry.objects.filter(
            entry_date__range=(m_start, m_end),
            entry_type='DEBIT',
            is_deleted=False
        ).aggregate(Sum('amount'))['amount__sum'] or 0.00
        
        m_credits = CashBookEntry.objects.filter(
            entry_date__range=(m_start, m_end),
            entry_type='CREDIT',
            is_deleted=False
        ).aggregate(Sum('amount'))['amount__sum'] or 0.00
        
        cashflow_chart.append({
            'label': f"{m}/{y}",
            'income': float(m_debits),
            'expense': float(m_credits)
        })
        
    # Expense category breakdown
    category_expenses = Expense.objects.filter(
        status='APPROVED',
        is_deleted=False
    ).values('category').annotate(total=Sum('amount')).order_by('-total')
    
    expense_breakdown = []
    for cat in category_expenses:
        expense_breakdown.append({
            'name': dict(Expense.CATEGORY_CHOICES).get(cat['category']),
            'value': float(cat['total'])
        })

    context = {
        'total_students': total_students,
        'total_staff': total_staff,
        'today_collections': today_collections,
        'monthly_collections': monthly_collections,
        'total_pending': total_pending,
        'monthly_expenses': monthly_expenses,
        'cash_in_hand': cash_in_hand,
        'bank_balance': bank_balance,
        'recent_transactions': recent_transactions,
        'upcoming_salaries': upcoming_salaries,
        'cashflow_chart': cashflow_chart,
        'expense_breakdown': expense_breakdown
    }
    return render(request, 'core/dashboard.html', context)

# Import datetime inside helper view if not global
import datetime

# Database Backup and Restore (SQLite based)
@login_required
@super_admin_only
def backup_restore_view(request):
    backup_dir = os.path.join(settings.BASE_DIR, 'backups')
    if not os.path.exists(backup_dir):
        os.makedirs(backup_dir)
        
    backups = sorted([f for f in os.listdir(backup_dir) if f.endswith('.sqlite3')], reverse=True)
    
    if request.method == 'POST' and 'create_backup' in request.POST:
        db_path = settings.DATABASES['default']['NAME']
        # check if it is sqlite database
        if 'sqlite3' in str(db_path):
            timestamp = now().strftime('%Y%m%d_%H%M%S')
            backup_file = os.path.join(backup_dir, f"backup_{timestamp}.sqlite3")
            shutil.copy2(db_path, backup_file)
            log_activity(request.user, "DATABASE_BACKUP", {"file": f"backup_{timestamp}.sqlite3"}, request)
            messages.success(request, f"Database backup 'backup_{timestamp}.sqlite3' successfully created.")
            return redirect('backup_restore')
        else:
            messages.error(request, "Database backup is only supported for SQLite configurations in this interface. For PostgreSQL, configure automated pg_dump cron jobs.")
            
    if request.method == 'POST' and 'restore_backup' in request.POST:
        backup_name = request.POST.get('backup_file')
        backup_file = os.path.join(backup_dir, backup_name)
        db_path = settings.DATABASES['default']['NAME']
        
        if os.path.exists(backup_file) and 'sqlite3' in str(db_path):
            # Overwrite current SQLite database
            shutil.copy2(backup_file, db_path)
            log_activity(request.user, "DATABASE_RESTORE", {"file": backup_name}, request)
            messages.success(request, f"Database successfully restored from '{backup_name}'.")
            return redirect('dashboard')
            
    return render(request, 'core/backup_restore.html', {'backups': backups})

# Reports & Exports Module
@login_required
@staff_only
def reports_dashboard(request):
    return render(request, 'core/reports_dashboard.html')

@login_required
@staff_only
def export_students_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Student Directory"
    
    # Styling
    title_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    cell_font = Font(name="Calibri", size=11)
    
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    title_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    
    # Title Row
    ws.merge_cells("A1:K1")
    ws["A1"] = "VYAS PUBLIC SCHOOL - STUDENT DIRECTORY"
    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 30
    
    # Headers
    headers = [
        "Admission No", "First Name", "Last Name", "Class-Section", "Roll No",
        "Guardian Name", "Mobile Number", "DOB", "Gender", "Category", "Status"
    ]
    ws.append([]) # blank spacer row
    ws.append(headers)
    ws.row_dimensions[3].height = 20
    
    # Format Headers
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        
    # Data Rows
    enrollments = StudentEnrollment.objects.filter(is_deleted=False).select_related(
        'student', 'class_section__class_level', 'class_section__section'
    ).order_by('class_section', 'roll_number')
    
    for en in enrollments:
        student = en.student
        ws.append([
            student.admission_number,
            student.first_name,
            student.last_name,
            str(en.class_section),
            en.roll_number or '-',
            student.guardian_name,
            student.guardian_mobile,
            student.date_of_birth.strftime('%Y-%m-%d'),
            student.get_gender_display(),
            student.get_category_display(),
            student.get_status_display()
        ])
        
    # Resize Columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = 'attachment; filename="student_directory.xlsx"'
    wb.save(response)
    return response

@login_required
@staff_only
def export_collections_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Collections Report"
    
    title_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    
    ws.merge_cells("A1:G1")
    ws["A1"] = "VYAS PUBLIC SCHOOL - FEE COLLECTIONS REPORT"
    ws["A1"].font = title_font
    ws["A1"].fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center")
    
    headers = ["Receipt No", "Date", "Admission No", "Student Name", "Class", "Amount Paid (₹)", "Payment Mode"]
    ws.append([])
    ws.append(headers)
    
    # Format Headers
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        
    collections = FeeCollection.objects.filter(is_deleted=False).select_related(
        'student_enrollment__student', 'student_enrollment__class_section__class_level'
    ).order_by('-payment_date', '-receipt_number')
    
    total_collected = 0.00
    for coll in collections:
        total_collected += float(coll.amount_paid)
        ws.append([
            coll.receipt_number,
            coll.payment_date.strftime('%Y-%m-%d'),
            coll.student_enrollment.student.admission_number,
            coll.student_enrollment.student.full_name,
            str(coll.student_enrollment.class_section),
            float(coll.amount_paid),
            coll.get_payment_mode_display()
        ])
        
    # Append Total
    ws.append([])
    ws.append(["", "", "", "", "Total Collection:", total_collected, ""])
    ws.cell(row=ws.max_row, column=5).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=6).font = Font(bold=True)
    
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = 'attachment; filename="fee_collections.xlsx"'
    wb.save(response)
    return response

def custom_permission_denied(request, exception=None):
    return render(request, '403.html', status=403)
